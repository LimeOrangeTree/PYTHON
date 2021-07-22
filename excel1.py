# import openpyxl
# wb=openpyxl.load_workbook('data\\address.xlsx')
# # print(type(wb))
# # print(wb.get_sheet_names())   #시트명
# # print(wb.sheetnames)
# # 시트에 접근
# # sheet=wb.get_sheet_by_name('friend')
# sheet=wb['friend']
# # print(type(sheet))
# # 현재 열린시트 확인
# sheet=wb.active
# # print(type(sheet))
# # 셀에 접근
# print(sheet['A3'])
# print(sheet['B3'].value)
# print(sheet.cell(row=3,column=3))
# print(sheet.cell(row=3,column=3).value)

# for i in range(1,10,2):
#     print(i,sheet.cell(row=i,column=2).value)

# colC=sheet['C']    #C열의 모든 데이터 출력
# for c in colC:
#     print(c.value)

# colCD=sheet['C:D']
# for c in colCD:
#     for cell in c:
#         print(cell.value)

# row59=sheet[5:9]
# for row in row59:
#     for cell in row:
#         print(cell.value)

# cellrange=sheet['a3','b4']
#-----엑셀문서 작성---------------------
# import openpyxl
# wb=openpyxl.Workbook()
# print(wb.sheetnames)
# sheet=wb.active
# print(sheet.title)   #시트이름 출력
# sheet.title='파이썬자동화'
# print(sheet.title)
# wb.create_sheet('python')  #마지막에 추가
# wb.create_sheet('mongoDB',1)  #지정된 위치에 추가
# sheet['a1']=300
# sheet['a2']=3
# sheet['a3']='=sum(a1:a2)'
#
# sheet.merge_cells('b2:d4')
# sheet['b2']='셀합치기'
#
# wb.save('data\\test1.xlsx')

# data1.xlsx 파일을 읽어서 Supplier Name 이 Supplier X가 아닌 데이터를
# data1.csv로 저장하세요===
# 1.파일을 읽을 범위를 결정
# 2. 행을 변경하면서 행데이터를 리스트로 만들어 저장
# import openpyxl
# def makeList():
#     wb=openpyxl.load_workbook('data\\data1.xlsx')
#     sheet=wb['data1']
#     result=[]
#     for row in sheet[1:13]:
#         # print(row)
#         if row[0].value.strip()!="Supplier X":
#             result.append(row)
#     return result
# def makeFile(result):
#     with open('data\\data1.csv','w') as f:
#         for items in result:   #[(),(),(),...]
#             str="{},{},{},{},{}\n".format\
#                 (items[0].value,items[1].value,items[2].value,items[3].value,items[4].value)
#             f.write(str)
# def main():
#     result=makeList()
#     makeFile(result)
# if __name__=="__main__":
#     main()
# stats.xlsx
import openpyxl

# 엑셀파일 열기
filename = "data\\stats.xlsx"
book = openpyxl.load_workbook(filename)

# 활성화된 시트 추출하기
sheet = book.active

# 서울 제외한 인구 구하기
for i in range(0,9):
    total = int(sheet[str(chr(i+66)) + "3"].value)
    seoul = int(sheet[str(chr(i+66)) + "4"].value)
    output = total - seoul
    print(" 서울 제외 인구: ", output)

    # 쓰기
    sheet[str(chr(i + 66)) + "21"] = output
    cell = sheet[str(chr(i + 66)) + "21"]

    # 폰트, 색상 변경해보기
    cell.font = openpyxl.styles.Font(size=14, color="F00000")
    # cell.number_format = cell.number_format

# 엑셀 파일 저장
filename = "data\\population.xlsx"
book.save(filename)
print("OK")
